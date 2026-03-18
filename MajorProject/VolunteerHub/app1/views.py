import json
from datetime import datetime
from django.conf import settings
from reportlab.pdfgen import canvas
from django.http import HttpResponse


from app1.models import Service
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.utils import timezone
from datetime import timedelta
from django.db.models import Count
from django.http import HttpResponse
import csv
from datetime import date, timedelta
from django.utils import timezone
from .models import Attendance
from .models import (
    Organization,
    Service,
    Application
)

User = get_user_model()


# ==================== AUTH ====================

@csrf_exempt
def login_page(request):
    return render(request, "login.html")


@csrf_exempt
def login_api(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=405)

    data = json.loads(request.body)
    email = data.get("username")
    password = data.get("password")
    role = data.get("role")

    user = User.objects.filter(email=email).first()
    if not user:
        return JsonResponse({"status": "error", "message": "Email not registered"})

    # ROLE VALIDATION
    if role == "VOLUNTEER":
        if not email.endswith("@srit.ac.in") or user.role != "VOLUNTEER":
            return JsonResponse({"status": "error", "message": "Invalid volunteer account"})
        if not VolunteerProfile.objects.filter(user=user).exists():
            return JsonResponse({"status": "error", "message": "Volunteer profile missing"})

    elif role == "ORGANIZATION":
        if not email.endswith("@gmail.com") or user.role != "ORGANIZATION":
            return JsonResponse({"status": "error", "message": "Invalid organization account"})
        org = Organization.objects.filter(user=user).first()
        if not org or not org.approved:
            return JsonResponse({"status": "error", "message": "Organization not approved by admin"})

    elif role == "ADMIN":
        if user.role != "ADMIN":
            return JsonResponse({"status": "error", "message": "Not an admin account"})

    else:
        return JsonResponse({"status": "error", "message": "Invalid role"})

    user = authenticate(request, username=user.username, password=password)

    if not user:
        return JsonResponse({"status": "error", "message": "Invalid password"})

    login(request, user)
    return JsonResponse({"status": "success", "role": user.role})


# ==================== HOME ====================

def home(request):
    return render(request, "home.html")


# ==================== DASHBOARDS ====================


from django.utils.timezone import now


@login_required
def volunteer_dashboard_page(request):
    if request.user.role != "VOLUNTEER":
        return redirect("login")

    profile, created = VolunteerProfile.objects.get_or_create(
        user=request.user,
        defaults={
            "full_name": request.user.email.split("@")[0],
            "phone": "",
            "student_id": "",
            "department": "",
            "year": "",
            "skills": ""
        }
    )

    today = now().date()

    # 🔢 COUNTS (THIS IS THE FIX)
    applied = Application.objects.filter(
        volunteer=profile,
        status="APPLIED"
    ).count()

    selected = Application.objects.filter(
        volunteer=profile,
        status="SELECTED"
    ).count()

    completed = Application.objects.filter(
        volunteer=profile,
        status="SELECTED",
        service__end_date__lt=today
    ).count()


    selected_apps = Application.objects.filter(
        volunteer=profile,
        status="SELECTED"
    ).select_related("service")

    total_days = 0
    total_present = 0

    today = timezone.now().date()

    for app in selected_apps:
        service = app.service

        if service.start_date and service.end_date:


            effective_end = min(service.end_date, today)

            if effective_end >= service.start_date:
                event_days = (effective_end - service.start_date).days + 1
                total_days += event_days

        present_days = Attendance.objects.filter(
            application=app,
            is_present=True
        ).count()

        total_present += present_days

    overall_attendance = (
        round((total_present / total_days) * 100, 2)
        if total_days > 0 else 0
    )
    ratings = Application.objects.filter(
        volunteer=profile,
        rating__gt=0  # only real ratings
    ).values_list("rating", flat=True)

    ratings = list(ratings)

    if ratings:
        avg_rating = round(sum(ratings) / len(ratings), 2)
        profile.rating = avg_rating
        profile.save()
    else:
        profile.rating = None
        avg_rating = None

    # 🎯 GOAL
    goal = 10

    # 📅 UPCOMING SERVICES
    # 📅 UPCOMING SERVICES (FIXED)

    upcoming_services = Application.objects.filter(
        volunteer=profile,
        status="SELECTED",
        service__start_date__gte=today
    ).select_related(
        "service",
        "service__organization",
        "service__organization__user"
    ).order_by("service__start_date")
    return render(
        request,
        "volunteer/dashboard.html",
        {
            "profile": profile,
            "applied_count": applied,
            "selected_count": selected,
            "completed": completed,
            "goal": goal,
            "upcoming_services": upcoming_services,
            "overall_attendance": overall_attendance,
            "active_page": "dashboard",
        }
    )


@login_required
def organization_dashboard_page(request):
    if request.user.role != "ORGANIZATION":
        return redirect("login")

    org = request.user.organization
    status_filter = request.GET.get("status")
    today = timezone.now().date()
    

    services = Service.objects.filter(organization=org)
    for service in services:
        if service.start_date <= today <= service.end_date:
            service.dynamic_status = "ACTIVE"
        elif today > service.end_date:
            service.dynamic_status = "CLOSED"
        else:
            service.dynamic_status = "UPCOMING"

    approved_count = services.filter(status="APPROVED").count()
    pending_count = services.filter(status="PENDING").count()
    status_filter = request.GET.get("status")
    if status_filter:
        services = [
            s for s in services
            if s.dynamic_status.lower() == status_filter.lower()
        ]

    

    # 🔥 Total Applicants Across All Works
    total_applicants = Application.objects.filter(
        service__organization=org
    ).count()

    # 🔥 Total Selected Volunteers
    total_selected = Application.objects.filter(
        service__organization=org,
        status="SELECTED"
    ).count()

    # 🔥 Total Completed Volunteers
    total_completed = Application.objects.filter(
        service__organization=org,
        status="COMPLETED"
    ).count()
    org = request.user.organization

    rejected_count = Service.objects.filter(
        organization=org,
        status="REJECTED"
    ).count()

    context = {
        "services": services,
        "approved_count": approved_count,
        "pending_count": pending_count,
        "total_applicants": total_applicants,
        "total_selected": total_selected,
        "total_completed": total_completed,
        "status_filter": status_filter,
        "today": today,
        "rejected_count": rejected_count,
    }

    return render(
        request,
        "organization/dashboard.html",
        context
    )

from datetime import date


@login_required
def admin_dashboard_page(request):

    if request.user.role != "ADMIN":
        return redirect("login")
    section = request.GET.get("section", "dashboard")
    

    today = date.today()

    volunteers = VolunteerProfile.objects.count()
    organizations = Organization.objects.filter(approved=True).count()

    # ✅ Active = approved + not expired
    active = Service.objects.filter(
        status__iexact="APPROVED",
        end_date__gte=today
    ).count()

    # ✅ Completed = expired by date
    completed = Service.objects.filter(
        end_date__lt=today
    ).count()

    pending_org_count = Organization.objects.filter(approved=False).count()
    pending_service_count = Service.objects.filter(status__iexact="PENDING").count()

    return render(
        request,
        "admin_panel/dashboard.html",
        {
            "volunteers": volunteers,
            "organizations": organizations,
            "active": active,
            "completed": completed,
            "pending_org_count": pending_org_count,
            "pending_service_count": pending_service_count,
        }
    )


@login_required
def admin_pending_services(request):
    if request.user.role != "ADMIN":
        return redirect("login")

    services = Service.objects.filter(status__iexact="pending")

    return render(
        request,
        "admin_panel/pending_services.html",
        {"services": services}
    )


@login_required
def admin_approve_service(request, service_id):
    if request.user.role != "ADMIN":
        return redirect("login")

    service = get_object_or_404(Service, id=service_id)
    service.status = "APPROVED"
    service.save()

    messages.success(request, "Service approved successfully")
    return redirect("admin_pending_services")

@login_required
def reject_service(request, service_id):

    if request.user.role != "ADMIN":
        return redirect("login")

    service = get_object_or_404(Service, id=service_id)
    service.status = "REJECTED"
    service.save()

    messages.success(request, "Service rejected successfully.")
    return redirect("admin_pending_services")
# ==================== ADMIN MODULE ====================
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import VolunteerProfile

@login_required
def admin_volunteers(request):
    if request.user.role != 'ADMIN':
        return redirect('login')

    volunteers = VolunteerProfile.objects.all()

    # ---- Normalize year ----
    def normalize_year(y):
        if "1" in y:
            return "1"
        if "2" in y:
            return "2"
        if "3" in y:
            return "3"
        if "4" in y:
            return "4"
        return None

    for v in volunteers:
        v.norm_year = normalize_year(v.year)

    # ---- Year counts ----
    year_counts = {
        "1": sum(1 for v in volunteers if v.norm_year == "1"),
        "2": sum(1 for v in volunteers if v.norm_year == "2"),
        "3": sum(1 for v in volunteers if v.norm_year == "3"),
        "4": sum(1 for v in volunteers if v.norm_year == "4"),
    }

    # ---- Branch counts (ONLY final year) ----
    branches = ["CSE", "ECE", "EEE", "MECH", "CIVIL"]

    branch_counts = {
        b: VolunteerProfile.objects.filter(department=b).count()
        for b in branches
    }

    # ---- Sorting ----
    sort = request.GET.get("sort", "roll")
    if sort == "name":
        volunteers = sorted(volunteers, key=lambda x: x.full_name.lower())
    else:
        volunteers = sorted(volunteers, key=lambda x: x.student_id)

    return render(request, "admin_panel/volunteers.html", {
        "volunteers": volunteers,
        "year_counts": year_counts,
        "branch_counts": branch_counts,
    })

@login_required
def admin_pending_organizations(request):
    if request.user.role != "ADMIN":
        return redirect("login")

    pending_orgs = Organization.objects.filter(approved=False)
    return render(request, "admin_panel/pending_organizations.html", {
        "organizations": pending_orgs
    })


@login_required
def approve_organization(request, org_id):
    if request.user.role != "ADMIN":
        return redirect("login")

    org = get_object_or_404(Organization, id=org_id)
    org.approved = True
    org.user.is_active = True
    org.user.save()
    org.save()

    messages.success(request, "Organization approved successfully")
    return redirect("/admin_panel/pending-organizations/")


@login_required
def reject_organization(request, org_id):
    if request.user.role != "ADMIN":
        return redirect("login")

    org = get_object_or_404(Organization, id=org_id)
    org.user.delete()

    messages.error(request, "Organization rejected")
    return redirect("/admin_panel/pending-organizations/")


# ==================== SERVICES ====================
@require_POST
@login_required

@login_required
def list_services(request):
    services = Service.objects.filter(status="APPROVED")
    return JsonResponse({"services": list(services.values())})


@require_POST
@login_required
def apply_service(request, service_id):
    if request.user.role != "VOLUNTEER":
        return JsonResponse({"error": "Unauthorized"}, status=403)

    profile = VolunteerProfile.objects.get(user=request.user)
    service = get_object_or_404(Service, id=service_id)

    Application.objects.create(volunteer=profile, service=service)
    return JsonResponse({"message": "Applied successfully"})


# ==================== AI SELECTION ====================

@require_POST
@login_required
def assign_volunteers(request, service_id):
    if request.user.role != "ADMIN":
        return JsonResponse({"error": "Unauthorized"}, status=403)

    service = Service.objects.get(id=service_id)
    applications = Application.objects.filter(
        service=service
    ).select_related("volunteer").order_by("-id")

    scored = [(calculate_score(app.volunteer, service), app) for app in applications]
    scored.sort(reverse=True, key=lambda x: x[0])

    selected = scored[:service.required_volunteers]
    for _, app in selected:
        app.status = "SELECTED"
        app.save()
        # Auto withdraw other overlapping applications
        Application.objects.filter(
            volunteer=app.volunteer,
            status__in=["APPLIED", "WAITLIST"],
            service__start_date__lte=app.service.end_date,
            service__end_date__gte=app.service.start_date
        ).exclude(id=app.id).update(status="REJECTED")

    service.status = "APPROVED"
    service.save()

    return JsonResponse({"message": "AI-based selection completed"})


def calculate_score(volunteer, service):
    score = 0
    score += volunteer.attendance * 0.4

    service_skills = service.description.lower()
    volunteer_skills = volunteer.skills.lower()
    match = sum(1 for word in volunteer_skills.split(",") if word.strip() in service_skills)
    score += match * 10 * 0.4

    score += volunteer.rating * 0.2
    return score


# ==================== SUBMISSION & RATING ====================

@require_POST
@login_required
def submit_work(request, application_id):
    if request.user.role != "VOLUNTEER":
        return JsonResponse({"error": "Unauthorized"}, status=403)

    data = json.loads(request.body)
    app = Application.objects.get(id=application_id)

    app.submission_text = data.get("submission")
    app.status = "COMPLETED"
    app.save()

    volunteer = app.volunteer
    volunteer.completed_services += 1
    volunteer.save()

    return JsonResponse({"message": "Work submitted successfully"})


@require_POST
@login_required
def rate_volunteer(request, application_id):
    if request.user.role != "ORGANIZATION":
        return JsonResponse({"error": "Unauthorized"}, status=403)

    data = json.loads(request.body)
    app = Application.objects.get(id=application_id)

    app.rating = data.get("rating")
    app.review = data.get("review")
    app.save()

    ratings = Application.objects.filter(
        volunteer=app.volunteer, rating__gt=0
    ).values_list("rating", flat=True)

    app.volunteer.rating = sum(ratings) / len(ratings)
    app.volunteer.save()

    return JsonResponse({"message": "Volunteer rated"})


# ==================== REGISTER ====================

def validate_password(password):
    if len(password) < 8:
        return "Password must be at least 8 characters"
    if not re.search(r"[A-Z]", password):
        return "Must contain uppercase letter"
    if not re.search(r"[a-z]", password):
        return "Must contain lowercase letter"
    if not re.search(r"\d", password):
        return "Must contain number"
    if not re.search(r"[@$!%*?&]", password):
        return "Must contain special character"
    return None


from django.db import transaction
from django.contrib import messages
import re


def register_view(request):
    if request.method != "POST":
        return render(request, "register.html")
    if not request.session.get("email_verified") and request.method == "POST":
        messages.error(request, "Please verify email first")
        return redirect("register")

    role = request.POST.get("role")
    full_name = request.POST.get("full_name")
    email = request.POST.get("email")
    phone = request.POST.get("phone")
    year = request.POST.get("year") if role == "VOLUNTEER" else ""
    gender = request.POST.get("gender")
    student_id = request.POST.get("student_id")
    department = request.POST.get("department")
    skills = request.POST.get("skills", "")
    password = request.POST.get("password")
    rpassword = request.POST.get("rpassword")
    principal_name = request.POST.get("principal_name")


    # ---------- VALIDATIONS ----------
    # password strength validation
    password_error = validate_password(password)

    if password_error:
        messages.error(request, password_error)
        return render(request, "register.html", {
            "form_data": request.POST,
            "selected_role": role,
            "email_verified": True
        })

    if password != rpassword:
        messages.warning(request, "Passwords do not match")
        return render(request, "register.html", {
            "form_data": request.POST,
            "selected_role": role,
            "email_verified": True
        })

    if role == "VOLUNTEER" and not email.endswith("@srit.ac.in"):
        messages.warning(request, "Volunteer email must end with @srit.ac.in")
        return render(request, "register.html", {
            "form_data": request.POST,
            "selected_role": role,
            "email_verified": True
        })
    if role == "VOLUNTEER":
        if VolunteerProfile.objects.filter(student_id=student_id).exists():
            messages.warning(request, "Student ID already registered")
            return render(request, "register.html", {
                "form_data": request.POST,
                "selected_role": role,
                "email_verified": True
            })
        rollno = request.POST.get("student_id").lower()

        pattern = r'^\d{2}4g\d{1}a\d{2}[a-zA-Z0-9]{2}$'

        if not re.match(pattern, rollno):
            messages.error(
                request,
                "Invalid roll number format. Example: 224G1A0530"
            )
            return render(request, "register.html", {
                "form_data": request.POST,
                "selected_role": role,
                "email_verified": True,
            })
        email_prefix = email.split("@")[0]

        if email_prefix != rollno:
            messages.error(
                request,
                "Email must match your roll number"
            )
            return render(request, "register.html", {
                "form_data": request.POST,
                "selected_role": role,
                "email_verified": True
            })

    if role == "ORGANIZATION" and not email.endswith("@gmail.com"):
        messages.warning(request, "Organization email must end with @gmail.com")
        return render(request, "register.html", {
            "form_data": request.POST,
            "selected_role": role,
            "email_verified": True
        })

    if User.objects.filter(username=email).exists():
        messages.warning(request, "Email already registered")
        return render(request, "register.html", {
            "form_data": request.POST,
            "selected_role": role,
            "email_verified": True
        })

    # ---------- CREATE USER ----------
    with transaction.atomic():
    
        user = User.objects.create_user(
            username=email,
            email=email,
            
            password=password,
            role=role
        )

        if role == "VOLUNTEER":
            VolunteerProfile.objects.update_or_create(
                user=user,
                defaults={
                    "full_name": full_name,
                    "phone": phone,
                    "student_id": student_id,
                    "department": department,
                    "year": year,
                    "skills": skills,
                    "gender": gender, 
                }
            )


        if role == "VOLUNTEER":
            VolunteerProfile.objects.update_or_create(
                user=user,
                defaults={
                    "full_name": full_name,
                    "phone": phone,
                    "student_id": student_id,
                    "department": department,
                    "year": year,
                    "skills": skills,
                    "gender": gender, 
                }
            )

            messages.success(request, "Volunteer registered successfully")

        else:  # ORGANIZATION
            user.is_active = False
            user.save()

            phone = request.POST.get("phone")
            letter = request.FILES.get("verification_letter")

            Organization.objects.update_or_create(
                user=user,
                defaults={
                    "organization_name": full_name,
                    "phone": phone,  # ✅ ADD THIS
                    "principal_name": principal_name,
                    "verification_letter": letter,
                    "approved": False
                }
            )
            messages.success(
                request,
                "Organization registered. Await admin approval."
            )

    return redirect("login")


def logout_view(request):
    logout(request)  # destroys session
    return redirect('login')

from datetime import date
@login_required
def admin_active_works(request):
    if request.user.role != "ADMIN":
        return redirect("login")
    today = date.today()

    services = Service.objects.filter(
        status="APPROVED",
        end_date__gte=today   # 🔥 only future or today
    )

    context = {
        "services": services,
        "active_count": services.count(),
        "org_count": services.values("organization").distinct().count(),
        "total_required": sum(s.max_volunteers for s in services),
    }

    return render(
        request,
        "admin_panel/active_works.html",
        context
    )

from django.utils import timezone
from datetime import date

@login_required
def admin_completed_works(request):
    if request.user.role != "ADMIN":
        return redirect("login")

    today = date.today()

    services = Service.objects.filter(end_date__lt=today)

    return render(
        request,
        "admin_panel/completed_works.html",
        {
            "services": services,
            "completed_count": services.count()
        }
    )
@login_required
@login_required
def organization_create_service(request):

    if request.user.role != "ORGANIZATION":
        return redirect("login")

    if request.method == "POST":

        title = request.POST.get("title")
        description = request.POST.get("description")
        location = request.POST.get("location")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")
        required_volunteers = request.POST.get("required_volunteers")
        required_gender = request.POST.get("required_gender")
        required_year = request.POST.get("required_year")

        org = Organization.objects.get(user=request.user)

        # 🔥 ADD THIS BLOCK — Collect coordinators
        coordinators_list = []

        for key in request.POST:
            if key.startswith("coord_name_"):
                index = key.split("_")[-1]

                name = request.POST.get(f"coord_name_{index}")
                phone = request.POST.get(f"coord_phone_{index}")

                if name and phone:
                    coordinators_list.append({
                        "name": name.strip(),
                        "phone": phone.strip()
                    })

        print("Coordinators received:", coordinators_list)

        # 🔥 SAVE coordinators here
        Service.objects.create(
            title=title,
            description=description,
            location=location,
            start_date=start_date,
            end_date=end_date,
            max_volunteers=required_volunteers,
            organization=org,
            authorization_letter=request.FILES.get("authorization_letter"),
            status="PENDING",
            required_gender=required_gender,
            required_year=required_year,
            coordinators=coordinators_list   # ✅ IMPORTANT
        )

        messages.success(
            request,
            "Service submitted successfully. Await admin approval."
        )

        return redirect("organization_dashboard")

    return render(request, "organization/create_service.html")


@login_required
def admin_assign_volunteers_page(request, service_id):
    if request.user.role != "ADMIN":
        return redirect("login")

    service = get_object_or_404(Service, id=service_id)
    applications = Application.objects.filter(service=service)

    return render(
        request,
        "admin_panel/assign_volunteers.html",
        {
            "service": service,
            "applications": applications
        }
    )


# ---- AI SCORING ----
def calculate_score(volunteer, service):
    score = 0
    score += volunteer.attendance * 0.4
    score += volunteer.rating * 0.6
    return score


# ---- ADMIN ASSIGN VOLUNTEERS ----
from django.views.decorators.http import require_POST


@require_POST
def assign_volunteers(request, service_id):
    if request.user.role != "ADMIN":
        return JsonResponse({"error": "Unauthorized"}, status=403)

    service = Service.objects.get(id=service_id)

    applications = Application.objects.filter(
        service=service,
        status="APPLIED"
    )

    scored = []
    for app in applications:
        score = calculate_score(app.volunteer, service)
        scored.append((score, app))

    scored.sort(reverse=True, key=lambda x: x[0])
    selected = scored[:service.required_volunteers]

    for _, app in selected:
        app.status = "SELECTED"
        app.save()
        # Auto withdraw other overlapping applications
        Application.objects.filter(
            volunteer=app.volunteer,
            status__in=["APPLIED", "WAITLIST"],
            service__start_date__lte=app.service.end_date,
            service__end_date__gte=app.service.start_date
        ).exclude(id=app.id).update(status="REJECTED")

    service.status = "APPROVED"  # ACTIVE
    service.save()

    return redirect("admin_active_works")


@login_required
@require_POST
def admin_mark_service_completed(request, service_id):
    if request.user.role != "ADMIN":
        return redirect("home")

    service = get_object_or_404(Service, id=service_id, status="APPROVED")

    service.status = "COMPLETED"
    service.save()

    return redirect("admin_completed_works")

@login_required
def volunteer_available_services(request):

    if request.user.role != "VOLUNTEER":
        return redirect("login")

    today = timezone.now().date()
    status_filter = request.GET.get("status")

    # Base queryset (only approved services)
    services = Service.objects.filter(status="APPROVED")

    # Apply filter safely
    if status_filter == "active":
        services = services.filter(end_date__gte=today)

    elif status_filter == "closed":
        services = services.filter(end_date__lt=today)

    profile = VolunteerProfile.objects.get(user=request.user)

    applications = Application.objects.filter(volunteer=profile)

    applied_dict = {
        app.service.id: app.status
        for app in applications
    }

    return render(
        request,
        "volunteer/available_services.html",
        {
            "services": services,
            "applied_dict": applied_dict,
            "today": today,
            "active_page": "services",
            "status_filter": status_filter,
        }
    )

@login_required
def volunteer_apply_service(request, service_id):
    if request.user.role != "VOLUNTEER":
        return redirect("login")

    profile = VolunteerProfile.objects.get(user=request.user)


    service = get_object_or_404(
        Service,
        id=service_id,
        status="APPROVED"
    )
    # Check if volunteer already SELECTED for overlapping service
    selected_conflict = Application.objects.filter(
        volunteer=profile,
        status="SELECTED",
        service__start_date__lte=service.end_date,
        service__end_date__gte=service.start_date
    ).exists()

    if selected_conflict:
        messages.error(
            request,
            "You are already selected for another service on this date."
        )
        return redirect("volunteer_available_services")

    # Prevent duplicate apply
    if Application.objects.filter(
            volunteer=profile,
            service=service
    ).exists():
        messages.warning(
            request,
            "You have already sent a request for this service"
        )
        return redirect("volunteer_available_services")

    Application.objects.create(
        volunteer=profile,
        service=service,
        status="APPLIED"
    )

    messages.success(
        request,
        "Request sent for approval"
    )

    return redirect("volunteer_available_services")

@login_required
def organization_view_applicants(request, service_id):

    service = get_object_or_404(Service, id=service_id)

    # All applications for this service
    applications = Application.objects.filter(service=service)

    selected_apps = applications.filter(status="SELECTED")
    rejected_apps = applications.filter(status="REJECTED")
    waitlist_apps = applications.filter(status="WAITLIST")
    applied_apps = applications.filter(status="APPLIED")
    selected_count = Application.objects.filter(
        service=service,
        status="SELECTED"
    ).count()

    remaining_seats = service.max_volunteers - selected_count


    if service.required_gender != "ANY":
        applied_apps = applied_apps.filter(
            volunteer__gender=service.required_gender
        )


    if service.required_year:
        applied_apps = applied_apps.filter(
            volunteer__year=service.required_year
        )

    today = timezone.now().date()
    event_last_day = service.end_date

    # ✅ attendance records for today for this service
    today_attendance = Attendance.objects.filter(
        application__service=service,
        date=today
    )

    # ✅ who is present today
    present_ids_today = set(
        today_attendance.filter(is_present=True)
        .values_list("application_id", flat=True)
    )

    # ✅ who is marked (present or absent)
    marked_ids_today = set(
        today_attendance.values_list("application_id", flat=True)
    )

    # ✅ check if any selected volunteer not yet marked today
    unmarked_exists = selected_apps.exclude(
        id__in=marked_ids_today
    ).exists()

    # ✅ event active check
    event_active = (
        service.start_date
        and service.end_date
        and service.start_date <= today <= service.end_date
    )
    # Check if today is last day
    is_last_day = today == service.end_date

    # Check if attendance for today is marked for all selected volunteers
    selected_count = Application.objects.filter(
        service=service,
        status="SELECTED"
    ).count()

    today_attendance_count = Attendance.objects.filter(
        application__service=service,
        date=today
    ).count()

    attendance_completed_today = selected_count == today_attendance_count

    # Final condition
    rating_open = is_last_day and attendance_completed_today

    context = {
        "service": service,
        "selected_apps": selected_apps,
        "rejected_apps": rejected_apps,
        "waitlist_apps": waitlist_apps,
        "applied_apps": applied_apps,
        "event_active": event_active,
        "today": today,
        "present_ids_today": present_ids_today,
        "marked_ids_today": marked_ids_today,
        "unmarked_exists": unmarked_exists,
        "rating_open": rating_open,
        "remaining_seats": remaining_seats,
    }

    return render(request, "organization/view_applicants.html", context)

@login_required
def org_approve_volunteer(request, app_id):
    if request.user.role != "ORGANIZATION":
        return redirect("login")

    app = get_object_or_404(Application, id=app_id)

    # Security: only owning org can approve
    if app.service.organization.user != request.user:
        return redirect("organization_dashboard")

    app.status = "SELECTED"
    app.save()
    # Auto withdraw other overlapping applications
    Application.objects.filter(
        volunteer=app.volunteer,
        status__in=["APPLIED", "WAITLIST"],
        service__start_date__lte=app.service.end_date,
        service__end_date__gte=app.service.start_date
    ).exclude(id=app.id).update(status="REJECTED")

    return redirect(
    f"/organization/service/{app.service.id}/applications/?tab=applied"
)


@login_required
def org_reject_volunteer(request, app_id):
    if request.user.role != "ORGANIZATION":
        return redirect("login")

    app = get_object_or_404(Application, id=app_id)

    if app.service.organization.user != request.user:
        return redirect("organization_dashboard")


    app.status = "REJECTED"
    app.save()
    Notification.objects.create(
        user=app.volunteer.user,
        title="❌ Application Update",
        message=f"Your application was not selected.",
        link="/volunteer/applications/"
    )

    return redirect(
    f"/organization/service/{app.service.id}/applications/?tab=applied"
)


@require_POST
@login_required
def org_select_volunteers(request, service_id):
    if request.user.role != "ORGANIZATION":
        return redirect("login")

    org = Organization.objects.get(user=request.user)
    service = get_object_or_404(Service, id=service_id, organization=org)

    # Get selected application IDs from form
    selected_ids = request.POST.getlist("selected_volunteers")

    if not selected_ids:
        messages.warning(
            request,
            "Please select at least one volunteer."
        )
        return redirect("org_view_applicants", service_id=service.id)

    # All applications for this service
    applications = Application.objects.filter(service=service)

    for app in applications:
        if str(app.id) in selected_ids:
            app.status = "SELECTED"
            app.save()
            # Auto withdraw other overlapping applications
            Application.objects.filter(
                volunteer=app.volunteer,
                status__in=["APPLIED", "WAITLIST"],
                service__start_date__lte=app.service.end_date,
                service__end_date__gte=app.service.start_date
            ).exclude(id=app.id).update(status="REJECTED")

        else:
            app.status = "REJECTED"
            app.save()


    # Mark service as ACTIVE (service is now running)
    service.status = "ACTIVE"
    service.save()

    messages.success(
        request,
        "Volunteers selected successfully. Service is now ACTIVE."
    )

    return redirect("organization_dashboard")


@login_required
def org_view_volunteer_profile(request, volunteer_id):
    if request.user.role != "ORGANIZATION":
        return redirect("login")

    volunteer = get_object_or_404(
        VolunteerProfile,
        id=volunteer_id
    )

    return render(
        request,
        "organization/volunteer_profile_view.html",
        {
            "volunteer": volunteer
        }
    )


@login_required
def volunteer_applications(request):
    if request.user.role != "VOLUNTEER":
        return redirect("login")

    profile = VolunteerProfile.objects.get(user=request.user)

    applications = Application.objects.filter(
        volunteer=profile
    ).select_related("service")

    return render(
        request,
        "volunteer/my_applications.html",
        {
            "applications": applications,
            "active_page": "applications"
        }
    )


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import VolunteerProfile


@login_required
def volunteer_profile(request):

    profile, created = VolunteerProfile.objects.get_or_create(
        user=request.user
    )

    if request.method == "POST":

        # Only update if field exists in POST
        if request.POST.get("full_name"):
            profile.full_name = request.POST.get("full_name")

        if request.POST.get("phone"):
            profile.phone = request.POST.get("phone")

        if request.POST.get("year"):
            profile.year = request.POST.get("year")

        # Skills
        skills = request.POST.get("skills")
        if skills is not None:
            profile.skills = skills

        # Images
        if request.FILES.get("photo"):
            profile.photo = request.FILES.get("photo")

        if request.FILES.get("cover_photo"):
            profile.cover_photo = request.FILES.get("cover_photo")

        profile.save()

        return redirect("/volunteer/profile/?updated=true")

    return render(
        request,
        "volunteer/profile.html",
        {"profile": profile}
    )

@login_required
def admin_approved_organizations(request):
    if request.user.role != "ADMIN":
        return redirect("login")

    organizations = Organization.objects.filter(approved=True)
    for org in organizations:
        org.service_count = Service.objects.filter(organization=org).count()

    return render(
        request,
        "admin_panel/approved_organizations.html",
        {
            "organizations": organizations,
            "active_page": "approved_orgs"
        }
    )

@login_required
def admin_organization_detail(request, org_id):
    if request.user.role != "ADMIN":
        return redirect("login")

    organization = get_object_or_404(Organization, id=org_id)

    return render(request, "admin_panel/organization_detail.html", {
        "organization": organization
    })

@require_POST
def mark_bulk_attendance(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    if request.method == "POST":


        today = timezone.now().date()

        present_ids = request.POST.getlist("present_ids")

        # all selected volunteers for this service
        selected_apps = Application.objects.filter(
            service=service,
            status="SELECTED"
        )

        for app in selected_apps:
            is_present = str(app.id) in present_ids

            Attendance.objects.update_or_create(
                application=app,
                date=today,
                defaults={
                    "is_present": is_present
                }
            )

    return redirect("org_view_applicants", service_id=service.id)

@login_required
def volunteer_attendance(request):

    profile = VolunteerProfile.objects.get(user=request.user)

    # All selected applications
    applications = Application.objects.filter(
        volunteer=profile,
        status="SELECTED"
    )

    attendance_data = []

    for app in applications:

        service = app.service

        total_days = (
            (service.end_date - service.start_date).days + 1
            if service.start_date and service.end_date
            else 0
        )

        present_days = Attendance.objects.filter(
            application=app,
            is_present=True
        ).count()

        percentage = (
            round((present_days / total_days) * 100, 2)
            if total_days > 0 else 0
        )

        attendance_data.append({
            "service": service,
            "total_days": total_days,
            "present_days": present_days,
            "percentage": percentage,
            "rating": app.rating
        })

    return render(request, "volunteer/attendance.html", {
        "attendance_data": attendance_data
    })


@login_required
def rate_volunteer(request, app_id):

    # 🔐 Only organization can rate
    if request.user.role != "ORGANIZATION":
        return JsonResponse({"success": False, "error": "Unauthorized"})

    app = get_object_or_404(Application, id=app_id)

    # 🔐 Only service owner can rate
    if app.service.organization.user != request.user:
        return JsonResponse({"success": False, "error": "Not allowed"})

    today = timezone.now().date()

    # ✅ Rating only on last day
    if today != app.service.end_date:
        return JsonResponse({"success": False, "error": "Not last day"})

    # ✅ Attendance must be fully completed for today
    selected_count = Application.objects.filter(
        service=app.service,
        status="SELECTED"
    ).count()

    today_attendance_count = Attendance.objects.filter(
        application__service=app.service,
        date=today
    ).count()

    if selected_count != today_attendance_count:
        return JsonResponse({"success": False, "error": "Attendance not completed"})

    # 🚫 Prevent re-rating
    if app.rating:
        return JsonResponse({"success": False, "error": "Already rated"})

    # ======================
    # ⭐ SAVE RATING
    # ======================
    if request.method == "POST":
        data = json.loads(request.body)
        rating = int(data.get("rating"))

        if rating < 1 or rating > 5:
            return JsonResponse({"success": False, "error": "Invalid rating"})

        app.rating = rating
        app.save()

        # ======================
        # 🔄 UPDATE VOLUNTEER AVG
        # ======================
        ratings = Application.objects.filter(
            volunteer=app.volunteer,
            rating__isnull=False
        ).values_list("rating", flat=True)

        avg = sum(ratings) / len(ratings)

        app.volunteer.rating = round(avg, 2)
        app.volunteer.save()

        return JsonResponse({
            "success": True,
            "new_average": app.volunteer.rating
        })

    return JsonResponse({"success": False})

from django.db.models import Avg, Count
from django.utils import timezone
from django.shortcuts import get_object_or_404, redirect

@login_required
@transaction.atomic
@require_POST
def auto_select_volunteers(request, service_id):

    if request.user.role != "ORGANIZATION":
        return redirect("login")

    service = get_object_or_404(Service, id=service_id)

    total_required = service.max_volunteers
    print("SERVICE ID:", service.id)
    print("MAX VOLUNTEERS:", service.max_volunteers)

    # ---------------------------------------
    # STEP 1: Reset ALL applications to APPLIED
    # (Important for clean recalculation)
    # ---------------------------------------
    Application.objects.filter(
        service=service
    ).update(status="APPLIED")

    # ---------------------------------------
    # STEP 2: Fetch fresh applied list
    # ---------------------------------------
    applications = Application.objects.filter(
        service=service,
        status="APPLIED"
    ).select_related("volunteer")

    total_applied = applications.count()

    if total_applied == 0:
        return redirect("org_view_applicants", service_id=service.id)

    # If applied <= required → select all
    if total_applied <= total_required:
        applications.update(status="SELECTED")
        return redirect("org_view_applicants", service_id=service.id)

    # ---------------------------------------
    # STEP 3: Separate Freshers & Experienced
    # ---------------------------------------
    experienced = []
    freshers = []

    for app in applications:

        completed_count = Application.objects.filter(
            volunteer=app.volunteer,
            status="COMPLETED"
        ).count()

        if completed_count > 0:
            experienced.append(app)
        else:
            freshers.append(app)

    exp_count = len(experienced)
    fresher_count = len(freshers)

    # ---------------------------------------
    # STEP 4: Dynamic Proportional Allocation
    # ---------------------------------------

    # Example:
    # 15 applied → 9 experienced, 6 freshers
    # Required = 10
    # Exp share = (9/15)*10 = 6
    # Fresher share = 4

    selected_exp = 0
    selected_fresher = 0

    if total_applied > 0:
        selected_exp = round((exp_count / total_applied) * total_required)
        selected_fresher = total_required - selected_exp

    # Do not exceed available
    selected_exp = min(selected_exp, exp_count)
    selected_fresher = min(selected_fresher, fresher_count)

    # ---------------------------------------
    # STEP 5: Scoring System
    # ---------------------------------------
    def calculate_score(app):

        volunteer = app.volunteer

        # Attendance %
        total_att = Attendance.objects.filter(
            application__volunteer=volunteer
        ).count()

        present_att = Attendance.objects.filter(
            application__volunteer=volunteer,
            is_present=True
        ).count()

        attendance_percentage = (
            (present_att / total_att) * 100
            if total_att > 0 else 0
        )

        rating = volunteer.rating or 0
        rating_scaled = rating * 20  # Convert 5 scale → 100 scale

        completed = Application.objects.filter(
            volunteer=volunteer,
            status="COMPLETED"
        ).count()

        # Experienced Scoring
        if completed > 0:
            return (attendance_percentage * 0.6) + (rating_scaled * 0.4)

        # Fresher Scoring
        year_bonus = 20 if volunteer.year == "4" else 15
        skill_bonus = 20 if volunteer.skills else 10
        return year_bonus + skill_bonus

    # Sort both groups
    experienced.sort(key=lambda x: calculate_score(x), reverse=True)
    freshers.sort(key=lambda x: calculate_score(x), reverse=True)

    # ---------------------------------------
    # STEP 6: Select Based on Allocation
    # ---------------------------------------
    selected_list = []

    selected_list.extend(experienced[:selected_exp])
    selected_list.extend(freshers[:selected_fresher])

    # ---------------------------------------
    # STEP 7: Fill Remaining Slots (Important)
    # ---------------------------------------
    if len(selected_list) < total_required:

        remaining_slots = total_required - len(selected_list)

        remaining_pool = (
            experienced[selected_exp:] +
            freshers[selected_fresher:]
        )

        remaining_pool.sort(
            key=lambda x: calculate_score(x),
            reverse=True
        )

        selected_list.extend(remaining_pool[:remaining_slots])

    # Final safety cut
    selected_list = selected_list[:total_required]

    selected_ids = [app.id for app in selected_list]

    # ---------------------------------------
    # STEP 8: Final Status Update
    # ---------------------------------------

        # Reject all first
    Application.objects.filter(service=service).update(status="REJECTED")

    # Select top required
    selected_ids = [app.id for app in selected_list]

    selected_apps = Application.objects.filter(id__in=selected_ids)

    for app in selected_apps:
        app.status = "SELECTED"
        app.save()
        # Auto withdraw other overlapping applications
        Application.objects.filter(
            volunteer=app.volunteer,
            status__in=["APPLIED", "WAITLIST"],
            service__start_date__lte=app.service.end_date,
            service__end_date__gte=app.service.start_date
        ).exclude(id=app.id).update(status="REJECTED")

    # Create waitlist from remaining high scores
    remaining_apps = [app for app in applications if app.id not in selected_ids]

    waitlist_count = 3   # 🔥 You can control this number

    waitlist_ids = [app.id for app in remaining_apps[:waitlist_count]]

    wait_apps = Application.objects.filter(id__in=waitlist_ids)

    for app in wait_apps:
        app.status = "WAITLIST"
        app.save()

        Notification.objects.create(
            user=app.volunteer.user,
            title="🤖 You Were in Waitlist!",
            message=f"You have been in Waitlist for {service.title}.",
            link="/volunteer/applications/"
        )

    print("TOTAL APPLIED:", total_applied)
    print("FRESHERS:", fresher_count)
    print("EXPERIENCED:", exp_count)
    print("FINAL SELECTED:", len(selected_ids))

    return redirect("org_view_applicants", service_id=service.id)

from datetime import date

@login_required
def organization_profile(request):

    if request.user.role != "ORGANIZATION":
        return redirect("login")

    organization = get_object_or_404(
        Organization,
        user=request.user
    )

    today = date.today()

    # 📊 Statistics
    total_services = Service.objects.filter(
        organization=organization
    ).count()

    active_services = Service.objects.filter(
        organization=organization,
        end_date__gte=today
    ).count()

    completed_services = Service.objects.filter(
        organization=organization,
        end_date__lt=today
    ).count()

    return render(
        request,
        "organization/profile.html",
        {
            "organization": organization,
            "total_services": total_services,
            "active_services": active_services,
            "completed_services": completed_services,
        }
    )
import re

@login_required
def edit_organization_profile(request):

    if request.user.role != "ORGANIZATION":
        return redirect("login")

    organization = get_object_or_404(
        Organization,
        user=request.user
    )

    def extract_youtube_id(url):
        if not url:
            return None

        if re.match(r'^[a-zA-Z0-9_-]{11}$', url):
            return url

        match = re.search(r'(v=|youtu\.be/)([a-zA-Z0-9_-]{11})', url)
        if match:
            return match.group(2)

        return None

    if request.method == "POST":
        organization.organization_name = request.POST.get("organization_name")
        organization.about = request.POST.get("about")
        organization.mission = request.POST.get("mission")
        organization.vision = request.POST.get("vision")

        organization.website = request.POST.get("website") or None

        # ✅ FIXED PART
        full_video_link = request.POST.get("promo_video")
        organization.promo_video = extract_youtube_id(full_video_link)

        if request.FILES.get("cover_image"):
            organization.cover_image = request.FILES.get("cover_image")

        if request.FILES.get("logo"):
            organization.logo = request.FILES.get("logo")

        organization.save()

        messages.success(request, "Profile updated successfully!")
        return redirect("organization_profile")

    return render(
        request,
        "organization/edit_profile.html",
        {"organization": organization}
    )

@login_required
def view_volunteer_profile(request, volunteer_id):

    if request.user.role != "ORGANIZATION":
        return redirect("login")

    volunteer = get_object_or_404(
        VolunteerProfile,
        id=volunteer_id
    )

    # Previous completed events count
    completed_count = Application.objects.filter(
        volunteer=volunteer,
        status="COMPLETED"
    ).count()

    # Attendance percentage
    total_att = Attendance.objects.filter(
        application__volunteer=volunteer
    ).count()

    present_att = Attendance.objects.filter(
        application__volunteer=volunteer,
        is_present=True
    ).count()

    attendance_percentage = (
        round((present_att / total_att) * 100, 2)
        if total_att > 0 else 0
    )

    context = {
        "volunteer": volunteer,
        "completed_count": completed_count,
        "attendance_percentage": attendance_percentage,
    }

    return render(
        request,
        "organization/volunteer_profile.html",
        context
    )
from django.shortcuts import get_object_or_404
from django.urls import reverse

@login_required
def promote_waitlist(request, app_id):

    # First get the application
    application = get_object_or_404(Application, id=app_id)
    service = application.service

    # Security check
    if service.organization.user != request.user:
        return redirect("organization_dashboard")

    # Count selected volunteers
    selected_count = Application.objects.filter(
        service=service,
        status="SELECTED"
    ).count()

    # Promote only if seats available
    if selected_count < service.max_volunteers:
        application.status = "SELECTED"
        application.save()
        Notification.objects.create(
            user=application.volunteer.user,
            title="🎉 Promoted from Waitlist!",
            message=f"You are now selected for {service.title}.",
            link="/volunteer/applications/"
        )

    # Redirect to Selected tab
    url = reverse("org_view_applicants", args=[service.id])
    return redirect(f"{url}?tab=selected")
@login_required
def manual_select_volunteer(request, app_id):

    if request.user.role != "ORGANIZATION":
        return redirect("login")

    application = get_object_or_404(Application, id=app_id)
    service = application.service

    # Security check
    if service.organization.user != request.user:
        return redirect("organization_dashboard")

    # Check max capacity
    selected_count = Application.objects.filter(
        service=service,
        status="SELECTED"
    ).count()

    if selected_count >= service.max_volunteers:
        return redirect("org_view_applicants", service_id=service.id)

    # Select this volunteer
    application.status = "SELECTED"
    application.save()
    Notification.objects.create(
        user=application.volunteer.user,
        title="🎉 Application Approved",
        message=f"You have been selected for {service.title}.",
        link="/volunteer/applications/"
    )
    

    return redirect("org_view_applicants", service_id=service.id)

@login_required
def waitlist_volunteer(request, app_id):

    application = get_object_or_404(Application, id=app_id)
    service = application.service

    if service.organization.user != request.user:
        return redirect("organization_dashboard")

    application.status = "WAITLIST"
    application.save()

    return redirect(
    f"/organization/service/{service.id}/applications/?tab=applied"
)

from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .models import Application

@login_required
def approve_volunteer(request, app_id):

    application = get_object_or_404(Application, id=app_id)
    service = application.service

    # Security check
    if service.organization.user != request.user:
        return redirect("organization_dashboard")

    # Count already selected
    selected_count = Application.objects.filter(
        service=service,
        status="SELECTED"
    ).count()

    # If seats available → select
    if selected_count < service.max_volunteers:
        application.status = "SELECTED"
        application.save()
        Notification.objects.create(
            user=application.volunteer.user,
            title="🎉 Application Approved",
            message=f"You have been selected for {service.title}.",
            link="/volunteer/applications/"
        )

    return redirect("org_view_applicants", service_id=service.id)

@login_required
def request_absence(request, app_id):

    application = get_object_or_404(
        Application,
        id=app_id,
        volunteer__user=request.user,
        status="SELECTED"
    )

    application.absence_requested = True
    application.save()
    print("ABSENCE SAVED FOR:", application.id, application.absence_requested)

    return redirect("volunteer_dashboard")

@login_required
def approve_absence(request, app_id):

    application = get_object_or_404(Application, id=app_id)
    service = application.service

    if service.organization.user != request.user:
        return redirect("organization_dashboard")

    # STEP 1 — Remove selected volunteer
    application.absence_approved = True
    application.status = "ABSENT"
    application.save()
    application.refresh_from_db()
    print("AFTER:", application.id, application.status)

    # STEP 2 — Count current selected volunteers
    selected_count = Application.objects.filter(
        service=service,
        status="SELECTED"
    ).count()

    required = service.max_volunteers

    # STEP 3 — If seats available, promote
    if selected_count < required:

        next_waitlisted = Application.objects.filter(
            service=service,
            status="WAITLIST"
        ).first()

        if next_waitlisted:
            next_waitlisted.status = "SELECTED"
            next_waitlisted.save()

    return redirect(
        reverse("org_view_applicants", args=[service.id]) + "?tab=selected"
    )

@login_required
def organization_service_detail(request, service_id):
    if request.user.role != "ORGANIZATION":
        return redirect("login")

    service = get_object_or_404(Service, id=service_id)

    # Security: only owner can view
    if service.organization.user != request.user:
        return redirect("organization_dashboard")

    return render(
        request,
        "organization/service_detail.html",
        {
            "service": service
        }
    )

from django.shortcuts import get_object_or_404
from datetime import date

def public_organization_profile(request, org_id):
    organization = get_object_or_404(Organization, id=org_id)

    today = date.today()

    total_services = Service.objects.filter(
        organization=organization
    ).count()

    active_services = Service.objects.filter(
        organization=organization,
        end_date__gte=today
    ).count()

    completed_services = Service.objects.filter(
        organization=organization,
        end_date__lt=today
    ).count()

    return render(request, "organization/public_profile.html", {
        "organization": organization,
        "total_services": total_services,
        "active_services": active_services,
        "completed_services": completed_services,
    })

import random
from django.core.mail import send_mail
from django.contrib.auth import get_user_model

User = get_user_model()

from django.http import JsonResponse
import random
from datetime import timedelta
from django.utils import timezone

def forgot_password(request):
    if request.method == "POST":
        email = request.POST.get("email")

        try:
            user = User.objects.get(email=email)

            otp = str(random.randint(100000, 999999))

            request.session["reset_email"] = email
            request.session["reset_otp"] = otp
            request.session["otp_created_at"] = timezone.now().isoformat()
            request.session["otp_verified"] = False

            send_otp_email(email, otp)

            return JsonResponse({"status": "success"})

        except User.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Email not registered"})

    return render(request, "auth/forgot_password.html")


from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags


def send_otp_email(user_email, otp):
    subject = "Your Password Reset Code | VolunteerHub"

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <body style="margin:0;padding:0;background-color:#f4f6f9;font-family:Arial,sans-serif;">
        <table width="100%" cellspacing="0" cellpadding="0" bgcolor="#f4f6f9">
            <tr>
                <td align="center">
                    <table width="500" cellpadding="0" cellspacing="0" style="background:#ffffff;margin-top:40px;border-radius:10px;overflow:hidden;box-shadow:0 5px 15px rgba(0,0,0,0.1);">

                        <!-- Header -->
                        <tr>
                            <td style="background:#1abc9c;padding:20px;text-align:center;color:#ffffff;">
                                <h2 style="margin:0;">VolunteerHub</h2>
                            </td>
                        </tr>

                        <!-- Body -->
                        <tr>
                            <td style="padding:30px;color:#333;">
                                <h3>Password Reset Request</h3>
                                <p>Hello,</p>
                                <p>We received a request to reset your password for your VolunteerHub account.</p>

                                <p style="text-align:center;margin:30px 0;">
                                    <span style="font-size:28px;font-weight:bold;letter-spacing:5px;background:#f4f6f9;padding:15px 25px;border-radius:8px;display:inline-block;">
                                        {otp}
                                    </span>
                                </p>

                                <p>This code will expire in <strong>5 minutes</strong>.</p>
                                <p>If you did not request this, please ignore this email or contact our support team immediately.</p>

                                <p style="margin-top:30px;">
                                    Regards,<br>
                                    <strong>VolunteerHub Security Team</strong>
                                </p>
                            </td>
                        </tr>

                        <!-- Footer -->
                        <tr>
                            <td style="background:#f4f6f9;padding:15px;text-align:center;font-size:12px;color:#777;">
                                © 2026 VolunteerHub. All rights reserved.
                            </td>
                        </tr>

                    </table>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """

    text_content = strip_tags(html_content)

    email = EmailMultiAlternatives(
        subject,
        text_content,
        settings.DEFAULT_FROM_EMAIL,
        [user_email],
    )
    email.attach_alternative(html_content, "text/html")
    email.send()
def verify_otp(request):
    if request.method == "POST":
        entered_otp = request.POST.get("otp")

        stored_otp = request.session.get("reset_otp")
        otp_created_at = request.session.get("otp_created_at")

        if not stored_otp:
            return JsonResponse({"status": "error", "message": "Session expired"})

        created_time = timezone.datetime.fromisoformat(otp_created_at)

        if timezone.now() > created_time + timedelta(minutes=5):
            return JsonResponse({"status": "error", "message": "OTP expired"})

        if entered_otp != stored_otp:
            return JsonResponse({"status": "error", "message": "Invalid OTP"})

        request.session["otp_verified"] = True
        return JsonResponse({"status": "success"})

    return JsonResponse({"status": "error"})
def reset_password(request):
    if request.method == "POST":
        if not request.session.get("otp_verified"):
            return JsonResponse({"status": "error", "message": "Unauthorized"})

        password = request.POST.get("password")
        email = request.session.get("reset_email")

        error = validate_password(password)
        if error:
            return JsonResponse({"status": "error", "message": error})

        user = User.objects.get(email=email)
        user.set_password(password)
        user.save()

        request.session.flush()

        return JsonResponse({"status": "success"})

    return JsonResponse({"status": "error"})


from django.http import JsonResponse
from .models import Notification

@login_required
def get_notifications(request):
    notifications = Notification.objects.filter(user=request.user)[:10]

    # 🔥 Delete notifications read more than 1 day ago
    Notification.objects.filter(
        is_read=True,
        read_at__lte=timezone.now() - timedelta(days=1)
    ).delete()

    notifications = Notification.objects.filter(
        user=request.user
    ).order_by("-created_at")

    unread_count = notifications.filter(is_read=False).count()
    data = []
    for n in notifications:
        data.append({
            "id": n.id,
            "title": n.title,
            "message": n.message,
            "link": n.link,
            "is_read": n.is_read,
            "created_at": n.created_at.strftime("%b %d, %H:%M")
        })

    unread_count = Notification.objects.filter(
        user=request.user,
        is_read=False
    ).count()

    return JsonResponse({
        "notifications": data,
        "unread_count": unread_count
    })


from django.utils import timezone

@login_required
def mark_notification_read(request, notif_id):

    notif = Notification.objects.get(id=notif_id, user=request.user)

    notif.is_read = True
    notif.read_at = timezone.now()   # ✅ save read time
    notif.save()

    return JsonResponse({"status": "ok"})
from django.utils import timezone
from datetime import timedelta, datetime
from django.shortcuts import render
from .models import Organization, Service, Application
from django.shortcuts import render
from django.utils import timezone
from datetime import timedelta, datetime
from .models import Organization, Service, Application


def reports_page(request):

    report_type = request.GET.get("type")
    today = timezone.now().date()

    start_date = None
    end_date = None

    # -------------------------
    # DATE FILTER LOGIC
    # -------------------------

    # DAY → Exactly 1 Day
    if report_type == "day":
        selected_date = request.GET.get("single_date")

        if selected_date:
            start_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
        else:
            start_date = today

        end_date = start_date


    # WEEK → Exactly 7 Days
    elif report_type == "week":
        selected_date = request.GET.get("single_date")

        if selected_date:
            end_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
        else:
            end_date = today

        start_date = end_date - timedelta(days=6)


    # YEAR → Full Year (365 / 366 auto handled)
    elif report_type == "year":
        selected_year = request.GET.get("year")

        if selected_year:
            year = int(selected_year)
        else:
            year = today.year

        start_date = datetime(year, 1, 1).date()
        end_date = datetime(year, 12, 31).date()


    # CUSTOM → Any Range
    elif report_type == "custom":
        start = request.GET.get("start_date")
        end = request.GET.get("end_date")

        if start and end:
            start_date = datetime.strptime(start, "%Y-%m-%d").date()
            end_date = datetime.strptime(end, "%Y-%m-%d").date()
        else:
            start_date = today
            end_date = today

    # First Page Load
    else:
        return render(request, "admin_panel/reports.html", {
            "report_data": [],
            "type": None
        })


    # -------------------------
    # FETCH DATA
    # -------------------------

    report_data = []

    organizations = Organization.objects.filter(approved=True)

    for org in organizations:

        services = Service.objects.filter(
            organization=org,
            created_at__date__range=(start_date, end_date)
        )

        applications = Application.objects.filter(
            service__in=services
        )

        report_data.append({
            "organization_name": org.organization_name,

            # SERVICES
            "total_services": services.count(),
            "approved_services": services.filter(status="APPROVED").count(),
            "rejected_services": services.filter(status="REJECTED").count(),
            "pending_services": services.filter(status="PENDING").count(),
            "closed_services": services.filter(status="COMPLETED").count(),

            # APPLICATIONS
            "total_applications": applications.count(),
            "selected": applications.filter(status="SELECTED").count(),
            "rejected": applications.filter(status="REJECTED").count(),
            "waitlist": applications.filter(status="WAITLIST").count(),
        })


    return render(request, "admin_panel/reports.html", {
        "report_data": report_data,
        "type": report_type,
        "start_date": start_date,
        "end_date": end_date,
    })
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.platypus import TableStyle
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta, datetime
from .models import Organization, Service, Application
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from django.utils import timezone
from datetime import timedelta, datetime
from .models import Organization, Service, Application


def download_report(request):

    report_type = request.GET.get("type")
    today = timezone.now().date()

    # -------- FILTER LOGIC --------

    if report_type == "day":
        selected_date = request.GET.get("single_date")
        start_date = datetime.strptime(selected_date, "%Y-%m-%d").date() if selected_date else today
        end_date = start_date

    elif report_type == "week":
        selected_date = request.GET.get("single_date")
        end_date = datetime.strptime(selected_date, "%Y-%m-%d").date() if selected_date else today
        start_date = end_date - timedelta(days=6)

    elif report_type == "year":
        selected_year = request.GET.get("year")
        year = int(selected_year) if selected_year else today.year
        start_date = datetime(year, 1, 1).date()
        end_date = datetime(year, 12, 31).date()

    elif report_type == "custom":
        start = request.GET.get("start_date")
        end = request.GET.get("end_date")
        if start and end:
            start_date = datetime.strptime(start, "%Y-%m-%d").date()
            end_date = datetime.strptime(end, "%Y-%m-%d").date()
        else:
            start_date = today
            end_date = today
    else:
        start_date = today
        end_date = today

    # -------- CREATE EXCEL --------

    wb = Workbook()
    ws = wb.active
    ws.title = "System Report"

    headers = [
        "Organization Name",
        "Total Services",
        "Approved Services",
        "Rejected Services",
        "Completed Services",
        "Pending Services",
        "Total Applications",
        "Selected Volunteers",
        "Rejected Volunteers",
        "Waitlist Volunteers"
    ]

    ws.append(headers)

    # Bold header
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    organizations = Organization.objects.filter(approved=True)

    for org in organizations:

        services = Service.objects.filter(
            organization=org,
            created_at__date__range=(start_date, end_date)
        )

        applications = Application.objects.filter(service__in=services)

        row = [
            org.organization_name,
            services.count(),
            services.filter(status="APPROVED").count(),
            services.filter(status="REJECTED").count(),
            services.filter(status="COMPLETED").count(),
            services.filter(status="PENDING").count(),
            applications.count(),
            applications.filter(status="SELECTED").count(),
            applications.filter(status="REJECTED").count(),
            applications.filter(status="WAITLIST").count(),
        ]

        ws.append(row)

    # Auto adjust column width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=VolunteerHub_Report.xlsx'

    wb.save(response)
    return response

def admin_reports(request):

    report_type = request.GET.get("report_type")

    today = timezone.now().date()

    if report_type == "daily":
        start_date = today
        end_date = today

    elif report_type == "weekly":
        start_date = today - timedelta(days=7)
        end_date = today

    elif report_type == "yearly":
        start_date = today.replace(month=1, day=1)
        end_date = today

    elif report_type == "custom":
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

    # SERVICES DATA
    services = Service.objects.filter(created_at__date__range=[start_date, end_date])

    approved_services = services.filter(status="APPROVED").count()
    rejected_services = services.filter(status="REJECTED").count()
    pending_services = services.filter(status="PENDING").count()

    # APPLICATION DATA
    applications = Application.objects.filter(service__in=services)

    selected = applications.filter(status="SELECTED").count()
    rejected = applications.filter(status="REJECTED").count()
    waitlist = applications.filter(status="WAITLIST").count()

    context = {
        "approved_services": approved_services,
        "rejected_services": rejected_services,
        "pending_services": pending_services,
        "selected": selected,
        "rejected": rejected,
        "waitlist": waitlist,
    }

    return render(request, "admin_panel/reports.html", context)
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from .models import Organization, Service, Application


def download_selected_volunteers(request, org_name):

    org = Organization.objects.get(organization_name=org_name)

    services = Service.objects.filter(
        organization=org,
        status="APPROVED"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Selected Volunteers"

    row = 1

    # 🔥 Main Title (Merged & Centered)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1).value = f"{org.organization_name} - Selected Volunteers Report"
    ws.cell(row=row, column=1).font = Font(size=14, bold=True)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 2

    for service in services:

        selected_apps = Application.objects.filter(
            service=service,
            status="SELECTED"
        )

        if selected_apps.exists():

            # 🔥 Service Heading
            ws.cell(row=row, column=1).value = f"Service: {service.title}"
            ws.cell(row=row, column=1).font = Font(size=12, bold=True)
            row += 1

            headers = ["Name", "Email", "Phone", "Year", "Department", "Skills"]

            # Header Row
            for col, header in enumerate(headers, start=1):
                ws.cell(row=row, column=col).value = header
                ws.cell(row=row, column=col).font = Font(bold=True)

            row += 1

            # Volunteer Data
            for app in selected_apps:

                volunteer = app.volunteer
                user = volunteer.user

                ws.cell(row=row, column=1).value = volunteer.full_name
                ws.cell(row=row, column=2).value = user.email
                ws.cell(row=row, column=3).value = volunteer.phone
                ws.cell(row=row, column=4).value = volunteer.year
                ws.cell(row=row, column=5).value = volunteer.department
                ws.cell(row=row, column=6).value = volunteer.skills

                row += 1

            row += 2  # Space between services

    # 🔥 Safe Auto Column Width (No MergedCell Error)
    for col in range(1, 7):
        max_length = 0
        column_letter = get_column_letter(col)

        for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col, max_col=col):
            for cell in row_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 3

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        f'attachment; filename="{org.organization_name}_Selected_Volunteers.xlsx"'
    )

    wb.save(response)

    return response

from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from .models import Application, Service


def download_selected_pdf(request, service_id):

    service = Service.objects.get(id=service_id)
    selected_apps = Application.objects.filter(service=service, status="SELECTED")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="selected_volunteers_{service.id}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)

    styles = getSampleStyleSheet()
    elements = []

    # Header
    elements.append(Paragraph(f"<b>{service.organization.organization_name}</b>", styles["Title"]))
    elements.append(Spacer(1,10))
    elements.append(Paragraph(f"Service: {service.title}", styles["Heading2"]))
    elements.append(Paragraph(f"Date: {service.start_date}", styles["Normal"]))
    elements.append(Spacer(1,20))

    # Table Data
    data = [["Name", "Year", "Roll Number", "Department","Number"]]

    for app in selected_apps:
        data.append([
            app.volunteer.full_name,
            app.volunteer.year,
            app.volunteer.student_id,
            app.volunteer.department,
            app.volunteer.phone
        ])

    table = Table(data)

    table.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.grey),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("GRID",(0,0),(-1,-1),1,colors.black)
    ]))

    elements.append(table)

    doc.build(elements)

    return response

import openpyxl
from django.http import HttpResponse


def download_selected_excel(request, service_id):

    service = Service.objects.get(id=service_id)
    selected_apps = Application.objects.filter(service=service, status="SELECTED")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Selected Volunteers"

    sheet.append(["Name", "Year", "Roll Number", "Department","Number"])

    for app in selected_apps:
        sheet.append([
            app.volunteer.full_name,
            app.volunteer.year,
            app.volunteer.student_id,
            app.volunteer.department,
            app.volunteer.phone,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = f'attachment; filename="selected_volunteers_{service.id}.xlsx"'

    workbook.save(response)

    return response

from django.shortcuts import render, get_object_or_404
from .models import VolunteerProfile


def public_volunteer_profile(request, volunteer_id):

    volunteer = get_object_or_404(VolunteerProfile, id=volunteer_id)

    context = {
        "volunteer": volunteer
    }

    return render(
        request,
        "volunteer/public_volunteer_profile.html",
        context
    )
def send_registration_otp_email(user_email, otp):

    subject = "Verify Your Email | VolunteerHub"

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <body style="margin:0;padding:0;background-color:#f4f6f9;font-family:Arial,sans-serif;">
        <table width="100%" cellspacing="0" cellpadding="0" bgcolor="#f4f6f9">
            <tr>
                <td align="center">
                    <table width="500" cellpadding="0" cellspacing="0" 
                    style="background:#ffffff;margin-top:40px;border-radius:10px;
                    overflow:hidden;box-shadow:0 5px 15px rgba(0,0,0,0.1);">

                        <tr>
                            <td style="background:#1abc9c;padding:20px;text-align:center;color:#ffffff;">
                                <h2 style="margin:0;">VolunteerHub</h2>
                            </td>
                        </tr>

                        <tr>
                            <td style="padding:30px;color:#333;">
                                <h3>Email Verification</h3>

                                <p>Hello,</p>

                                <p>Thank you for registering with <strong>VolunteerHub</strong>.
                                Please use the OTP below to verify your email address.</p>

                                <p style="text-align:center;margin:30px 0;">
                                    <span style="font-size:28px;font-weight:bold;
                                    letter-spacing:5px;background:#f4f6f9;
                                    padding:15px 25px;border-radius:8px;
                                    display:inline-block;">
                                        {otp}
                                    </span>
                                </p>

                                <p>This OTP will expire in <strong>5 minutes</strong>.</p>

                                <p>If you did not create this account, please ignore this email.</p>

                                <p style="margin-top:30px;">
                                    Regards,<br>
                                    <strong>VolunteerHub Team</strong>
                                </p>
                            </td>
                        </tr>

                        <tr>
                            <td style="background:#f4f6f9;padding:15px;text-align:center;font-size:12px;color:#777;">
                                © 2026 VolunteerHub. All rights reserved.
                            </td>
                        </tr>

                    </table>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """

    text_content = strip_tags(html_content)

    email = EmailMultiAlternatives(
        subject,
        text_content,
        settings.DEFAULT_FROM_EMAIL,
        [user_email],
    )

    email.attach_alternative(html_content, "text/html")
    email.send()

import json
import random
from django.http import JsonResponse
from django.contrib.auth import get_user_model

User = get_user_model()

def send_register_otp(request):

    data = json.loads(request.body)

    email = data.get("email")
    role = data.get("role")

    # Volunteer domain check
    if role == "VOLUNTEER" and not email.endswith("@srit.ac.in"):
        return JsonResponse({
            "status":"error",
            "message":"Volunteer email must end with @srit.ac.in"
        })

    # Organization domain check
    if role == "ORGANIZATION" and not email.endswith("@gmail.com"):
        return JsonResponse({
            "status":"error",
            "message":"Organization email must end with @gmail.com"
        })

    # Email already exists
    if User.objects.filter(username=email).exists():
        return JsonResponse({
            "status":"error",
            "message":"Email already registered"
        })

    otp = str(random.randint(100000,999999))

    request.session["register_otp"] = otp
    request.session["register_email"] = email
    request.session["otp_created_at"] = timezone.now().isoformat()
    send_registration_otp_email(email, otp)

    return JsonResponse({
        "status":"success"
    })

import json
from datetime import timedelta

def verify_register_otp(request):

    data = json.loads(request.body)

    entered_otp = data.get("otp")

    stored_otp = request.session.get("register_otp")
    created_time = request.session.get("otp_created_at")
    if not stored_otp or not created_time:
        return JsonResponse({
            "status": "error",
            "message": "Session expired. Request OTP again."
        })
    created_time = timezone.datetime.fromisoformat(created_time)

    if timezone.now() > created_time + timedelta(minutes=5):
        return JsonResponse({
            "status":"error",
            "message":"OTP expired"
        })

    if entered_otp != stored_otp:
        return JsonResponse({
            "status":"error",
            "message":"Invalid OTP"
        })

    request.session["email_verified"] = True

    return JsonResponse({
        "status":"success"
    })